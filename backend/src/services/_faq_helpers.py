"""FAQ Service 内部辅助：embed/upsert、Excel 模板/解析/批量写入。

下划线前缀表明该模块是 ``services.faq_service`` 的私有实现细节，
外部模块（api/routes、其他 service、core）不应 import 本文件。

无 FastAPI 依赖：Excel 序列化只产出 bytes，HTTP 响应在路由层完成。
"""

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import get_config
from src.core.rag.embedding import get_embed_model
from src.storage.vector_store import VectorStore

logger = logging.getLogger(__name__)


# ── embed / upsert ────────────────────────────────────────────


def embed_faq_text(question: str, answer: str) -> tuple[list[float], str]:
    """将 Q+A 合并后 embed，返回 (vector, combined_text)。"""
    text = f"Q: {question}\nA: {answer}"
    embed_model = get_embed_model(text_type="document")
    vector = embed_model.get_text_embedding(text)
    return vector, text


def upsert_faq_vector(
    kb_name: str,
    question: str,
    answer: str,
    faq_id: int,
    vector_id: str,
    vec_store: VectorStore,
) -> None:
    """同步：embed 并 upsert 到 Qdrant（供 asyncio.to_thread 调用）。"""
    vector, text = embed_faq_text(question, answer)
    vec_store.add_vectors(
        collection_name=kb_name,
        vectors=[vector],
        payloads=[
            {
                "text": text,
                "source_type": "faq",
                "faq_id": faq_id,
                "source_file": "FAQ",
                "kb_name": kb_name,
            }
        ],
        ids=[vector_id],
    )


# ── Excel 工作簿 ──────────────────────────────────────────────


def build_faq_workbook(faqs: list[dict] | None = None) -> Workbook:
    """构建 FAQ Excel 工作簿。

    Args:
        faqs: None → 模板（含示例行）；list → 导出（含数据）。

    Returns:
        openpyxl Workbook 实例。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ"

    is_export = faqs is not None
    base_headers = ["问题 (必填)", "答案 (必填)", "分类", "排序号"]
    extra_headers = ["ID", "是否启用", "创建时间"]
    headers = base_headers + (extra_headers if is_export else [])

    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[1].height = 24

    col_widths = [40, 60, 15, 10] + ([8, 8, 20] if is_export else [])
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sample_fill = PatternFill(start_color="F8F6F2", end_color="F8F6F2", fill_type="solid")

    if is_export and faqs:
        for row_idx, faq in enumerate(faqs, 2):
            values = [
                faq["question"],
                faq["answer"],
                faq.get("category", ""),
                faq.get("sort_order", 0),
                faq["id"],
                "是" if faq.get("enabled") else "否",
                (faq["created_at"][:10] if faq.get("created_at") else ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 2:
                    cell.alignment = wrap_align
                    ws.row_dimensions[row_idx].height = 60
    elif not is_export:
        sample = [
            "郑州大学毕业答辩的时间安排是什么？",
            "郑州大学本科毕业论文答辩通常安排在每年6月初，具体时间由各学院通知，请关注学院官网。",
            "毕业答辩",
            0,
        ]
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = sample_fill
            if col_idx == 2:
                cell.alignment = wrap_align

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """将 openpyxl Workbook 序列化为字节。"""
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


# ── Excel 解析 ─────────────────────────────────────────────────


def parse_faq_sheet(ws) -> tuple[list[dict], list[dict]]:
    """解析 FAQ 工作表，返回 (valid_rows, errors)。

    Args:
        ws: openpyxl Worksheet 对象。

    Returns:
        valid_rows: [{"question", "answer", "category", "sort_order"}, ...]
        errors: [{"row", "question", "reason"}, ...]
    """
    valid_rows: list[dict] = []
    errors: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in (row[:4] if len(row) >= 4 else row)):
            continue

        question = str(row[0]).strip() if row[0] is not None else ""
        answer = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        category = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        sort_raw = row[3] if len(row) > 3 else None

        err: str | None = None
        sort_order = 0
        if not question:
            err = "问题不能为空"
        elif len(question) > 500:
            err = "问题超过 500 字符限制"
        elif not answer:
            err = "答案不能为空"
        elif len(answer) > 2000:
            err = "答案超过 2000 字符限制"
        elif category and len(category) > 64:
            err = "分类超过 64 字符限制"
        else:
            if sort_raw is not None and str(sort_raw).strip() != "":
                try:
                    sort_order = int(sort_raw)
                    if sort_order < 0:
                        err = "排序号不能为负数"
                except (ValueError, TypeError):
                    err = f"排序号 '{sort_raw}' 不是有效整数"

        if err:
            errors.append({"row": row_idx, "question": question[:50], "reason": err})
        else:
            valid_rows.append(
                {
                    "question": question,
                    "answer": answer,
                    "category": category,
                    "sort_order": sort_order,
                }
            )

    return valid_rows, errors


# ── 批量 embed ─────────────────────────────────────────────────


def batch_embed_and_upsert(
    kb_name: str,
    rows: list[dict],
    vec_store: VectorStore,
) -> dict[int, str | Exception]:
    """批量 embed 并 upsert 到 Qdrant。

    Args:
        kb_name: 知识库名称。
        rows: 每项包含 question, answer, faq_id, vector_id。
        vec_store: Qdrant 向量存储实例。

    Returns:
        {faq_id: vector_id} 或 {faq_id: Exception}。
    """
    cfg = get_config()
    batch_size: int = cfg.get("embedding", {}).get("embed_batch_size", 10)
    embed_model = get_embed_model(text_type="document")
    results: dict[int, str | Exception] = {}

    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        texts = [f"Q: {r['question']}\nA: {r['answer']}" for r in batch]
        try:
            vectors = embed_model.get_text_embedding_batch(texts)
            payloads = [
                {
                    "text": texts[j],
                    "source_type": "faq",
                    "faq_id": batch[j]["faq_id"],
                    "source_file": "FAQ",
                    "kb_name": kb_name,
                }
                for j in range(len(batch))
            ]
            vec_store.add_vectors(
                collection_name=kb_name,
                vectors=vectors,
                payloads=payloads,
                ids=[r["vector_id"] for r in batch],
            )
            for r in batch:
                results[r["faq_id"]] = r["vector_id"]
        except Exception as e:
            logger.warning("[faq import] batch %d embed/upsert 失败: %s", i // batch_size, e)
            for r in batch:
                results[r["faq_id"]] = e

    return results
