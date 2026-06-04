"""用户 Excel 批量导入 Mixin：学生 / 教师 / 师生关系。"""

import io
import logging
import zipfile

import pymysql

from src.exceptions import StorageError
from src.storage.interfaces.user_store import BaseUserStore


class ImportMixin:
    """用户 Excel 批量导入方法集合。"""

    _user_store: BaseUserStore
    logger: logging.Logger

    def import_students_from_xlsx(
        self,
        file_bytes: bytes,
        default_password: str,
        password_hasher: "callable[[str], str]",
        random_pwd_fn: "callable[[], str]",
    ) -> dict:
        """从 Excel 字节流批量导入学生账号。

        Args:
            file_bytes: Excel 文件的原始字节内容。
            default_password: 密码列为空时使用的默认密码，仍为空则由 random_pwd_fn 生成。
            password_hasher: 将明文密码转换为哈希值的函数。
            random_pwd_fn: 生成随机密码的函数。

        Returns:
            含 total / success / skipped / failed / errors 字段的汇总 dict。

        Raises:
            ValueError: Excel 文件格式无效或解析失败。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, skipped, failed = 0, 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            display_name = str(row[0]).strip() if row[0] else ""
            student_id = str(row[1]).strip() if row[1] else ""
            grade = str(row[2]).strip() if row[2] else ""
            major = str(row[3]).strip() if row[3] else ""
            class_name = str(row[4]).strip() if row[4] else ""
            password = str(row[5]).strip() if len(row) > 5 and row[5] else (default_password or random_pwd_fn())

            if not student_id:
                failed.append({"row": row_idx, "student_id": student_id, "reason": "学号不能为空"})
                continue

            if self._user_store.get_user_by_student_id(student_id):
                skipped += 1
                continue

            try:
                user = self._user_store.create_user(
                    username=student_id,
                    hashed_pwd=password_hasher(password),
                    display_name=display_name,
                    role="student",
                )
                self._user_store.upsert_student_profile(user["id"], student_id, grade, major, class_name)
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] student import row %d failed: %s", row_idx, e)
                failed.append({"row": row_idx, "student_id": student_id, "reason": str(e)})

        return {
            "total": success + skipped + len(failed),
            "success": success,
            "skipped": skipped,
            "failed": len(failed),
            "errors": failed,
        }

    def import_teachers_from_xlsx(
        self,
        file_bytes: bytes,
        default_password: str,
        password_hasher: "callable[[str], str]",
        random_pwd_fn: "callable[[], str]",
    ) -> dict:
        """从 Excel 字节流批量导入教师账号。

        Args 与 ``import_students_from_xlsx`` 一致。

        Returns:
            含 total / success / skipped / failed / errors 字段的汇总 dict。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, skipped, failed = 0, 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            display_name, emp_id, dept, title = [str(x).strip() if x else "" for x in row[:4]]
            password = str(row[4]).strip() if len(row) > 4 and row[4] else (default_password or random_pwd_fn())

            if not emp_id:
                failed.append({"row": row_idx, "employee_id": emp_id, "reason": "工号不能为空"})
                continue

            if self._user_store.get_user_by_employee_id(emp_id):
                skipped += 1
                continue

            try:
                user = self._user_store.create_user(emp_id, password_hasher(password), display_name, "teacher")
                self._user_store.upsert_teacher_profile(user["id"], emp_id, dept, title)
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] teacher import row %d failed: %s", row_idx, e)
                failed.append({"row": row_idx, "employee_id": emp_id, "reason": str(e)})

        return {
            "total": success + skipped + len(failed),
            "success": success,
            "skipped": skipped,
            "failed": len(failed),
            "errors": failed,
        }

    def import_mentor_relations_from_xlsx(self, file_bytes: bytes) -> dict:
        """从 Excel 字节流批量导入师生关系。

        Args:
            file_bytes: Excel 文件的原始字节内容（学号列 + 工号列）。

        Returns:
            含 total / success / failed / errors 字段的汇总 dict。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, failed = 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            student_id = str(row[0]).strip() if row[0] else ""
            employee_id = str(row[1]).strip() if row[1] else ""

            if not student_id or not employee_id:
                failed.append({"row": row_idx, "reason": "学号和工号均不能为空"})
                continue

            student_user = self._user_store.get_user_by_student_id(student_id)
            if not student_user:
                failed.append({"row": row_idx, "student_id": student_id, "reason": "学生不存在"})
                continue

            mentor_user = self._user_store.get_user_by_employee_id(employee_id)
            if not mentor_user or mentor_user["role"] != "teacher":
                failed.append(
                    {
                        "row": row_idx,
                        "employee_id": employee_id,
                        "reason": "导师不存在或非教师角色",
                    }
                )
                continue

            try:
                self._user_store.add_mentor_relation(mentor_user["id"], student_user["id"])
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] relation import row %d failed: %s", row_idx, e)
                failed.append(
                    {
                        "row": row_idx,
                        "student_id": student_id,
                        "employee_id": employee_id,
                        "reason": str(e),
                    }
                )

        return {
            "total": success + len(failed),
            "success": success,
            "failed": len(failed),
            "errors": failed,
        }
