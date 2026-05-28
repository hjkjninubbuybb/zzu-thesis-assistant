"""用户账号业务逻辑服务层。

职责：封装用户 CRUD、档案管理、师生关系、Excel 批量导入的全部业务逻辑。
本模块不包含任何 FastAPI / HTTP 相关导入。
"""

import io
import logging
import zipfile

import pymysql

from src.exceptions import StorageError, UserNotFoundError
from src.services.base import BaseService
from src.storage.interfaces.user_store import BaseUserStore

logger = logging.getLogger(__name__)


class UserService(BaseService):
    """用户账号增删查改及批量导入服务。"""

    def __init__(self, user_store: BaseUserStore) -> None:
        super().__init__()
        self._user_store = user_store

    # ── 用户基本操作 ──────────────────────────────────────────

    def list_users(
        self,
        role: str | None,
        page: int,
        page_size: int,
    ) -> tuple[list[dict], int]:
        """分页列出用户。

        Args:
            role: 角色过滤（admin / teacher / student），None 表示不过滤。
            page: 页码（从 1 开始）。
            page_size: 每页条数。

        Returns:
            (用户列表, 总数) 元组。
        """
        return self._user_store.list_users(role=role, page=page, page_size=page_size)

    def get_user(self, user_id: int) -> dict:
        """获取用户，不存在则抛出异常。

        Args:
            user_id: 用户 ID。

        Returns:
            用户 dict。

        Raises:
            UserNotFoundError: 用户不存在。
        """
        user = self._user_store.get_user_by_id(user_id)
        if not user:
            raise UserNotFoundError(f"用户 {user_id} 不存在")
        return user

    def get_user_by_username(self, username: str) -> dict | None:
        """按用户名查询用户。

        Args:
            username: 用户名。

        Returns:
            用户 dict，不存在返回 None。
        """
        return self._user_store.get_user_by_username(username)

    def create_user(
        self,
        username: str,
        hashed_pwd: str,
        display_name: str,
        role: str,
    ) -> dict:
        """创建新用户。

        Args:
            username: 用户名（唯一）。
            hashed_pwd: 已哈希的密码。
            display_name: 展示名称。
            role: 角色（admin / teacher / student）。

        Returns:
            新建用户 dict。
        """
        return self._user_store.create_user(
            username=username,
            hashed_pwd=hashed_pwd,
            display_name=display_name,
            role=role,
        )

    def update_user(self, user_id: int, **kwargs: object) -> dict:
        """更新用户字段，不存在则抛出异常。

        Args:
            user_id: 用户 ID。
            **kwargs: 可更新的字段（display_name / hashed_pwd / role / is_active）。

        Returns:
            更新后的用户 dict。

        Raises:
            UserNotFoundError: 用户不存在。
        """
        user = self._user_store.get_user_by_id(user_id)
        if not user:
            raise UserNotFoundError(f"用户 {user_id} 不存在")
        updated = self._user_store.update_user(user_id, **kwargs)
        return updated or user

    def delete_user(self, user_id: int) -> None:
        """删除用户，不存在则抛出异常。

        Args:
            user_id: 用户 ID。

        Raises:
            UserNotFoundError: 用户不存在。
        """
        user = self._user_store.get_user_by_id(user_id)
        if not user:
            raise UserNotFoundError(f"用户 {user_id} 不存在")
        self._user_store.delete_user(user_id)

    def add_login_log(self, user_id: int, ip_addr: str = "", user_agent: str = "") -> None:
        """记录用户登录日志（失败时静默忽略）。

        Args:
            user_id: 用户 ID。
            ip_addr: 客户端 IP 地址。
            user_agent: User-Agent 字符串。
        """
        try:
            self._user_store.add_login_log(user_id, ip_addr, user_agent)
        except Exception as e:
            logger.warning("[UserService] 记录登录日志失败: %s", e)

    # ── 档案操作 ──────────────────────────────────────────────

    def get_profile(self, user: dict) -> dict | None:
        """根据用户角色获取对应档案。

        Args:
            user: 用户 dict，需包含 id 和 role 字段。

        Returns:
            学生或教师档案 dict，无档案时返回 None。
        """
        role = user["role"]
        uid = user["id"]
        if role == "student":
            return self._user_store.get_student_profile(uid)
        if role in ("teacher", "admin"):
            return self._user_store.get_teacher_profile(uid)
        return None

    def upsert_student_profile(
        self,
        user_id: int,
        student_id: str,
        grade: str,
        major: str,
        class_name: str,
    ) -> dict:
        """创建或更新学生档案。

        Args:
            user_id: 用户 ID。
            student_id: 学号。
            grade: 年级。
            major: 专业。
            class_name: 班级。

        Returns:
            更新后的学生档案 dict。
        """
        return self._user_store.upsert_student_profile(user_id, student_id, grade, major, class_name)

    def upsert_teacher_profile(
        self,
        user_id: int,
        employee_id: str,
        department: str,
        title: str,
    ) -> dict:
        """创建或更新教师档案。

        Args:
            user_id: 用户 ID。
            employee_id: 工号。
            department: 院系。
            title: 职称。

        Returns:
            更新后的教师档案 dict。
        """
        return self._user_store.upsert_teacher_profile(user_id, employee_id, department, title)

    # ── 师生关系 ──────────────────────────────────────────────

    def list_mentor_students(self, mentor_id: int) -> list[dict]:
        """列出导师名下的学生。

        Args:
            mentor_id: 导师用户 ID。

        Returns:
            学生 dict 列表（含学生档案字段）。
        """
        return self._user_store.list_mentor_students(mentor_id)

    def get_student_mentor(self, student_id: int) -> dict | None:
        """获取学生的指导教师。

        Args:
            student_id: 学生用户 ID。

        Returns:
            教师 dict（含教师档案字段），未分配则返回 None。
        """
        return self._user_store.get_student_mentor(student_id)

    def add_mentor_relation(self, mentor_id: int, student_id: int) -> None:
        """建立师生关系。

        Args:
            mentor_id: 导师用户 ID。
            student_id: 学生用户 ID。
        """
        self._user_store.add_mentor_relation(mentor_id, student_id)

    def remove_mentor_relation(self, mentor_id: int, student_id: int) -> None:
        """解除师生关系。

        Args:
            mentor_id: 导师用户 ID。
            student_id: 学生用户 ID。
        """
        self._user_store.remove_mentor_relation(mentor_id, student_id)

    # ── 用户查找辅助 ──────────────────────────────────────────

    def get_user_by_student_id(self, student_id: str) -> dict | None:
        """通过学号查用户。

        Args:
            student_id: 学号字符串。

        Returns:
            用户 dict，不存在返回 None。
        """
        return self._user_store.get_user_by_student_id(student_id)

    def get_user_by_employee_id(self, employee_id: str) -> dict | None:
        """通过工号查用户。

        Args:
            employee_id: 工号字符串。

        Returns:
            用户 dict，不存在返回 None。
        """
        return self._user_store.get_user_by_employee_id(employee_id)

    # ── Excel 批量导入 ────────────────────────────────────────

    def import_students_from_xlsx(
        self,
        file_bytes: bytes,
        default_password: str,
        password_hasher: "callable[[str], str]",
        random_pwd_fn: "callable[[], str]",
    ) -> dict:
        """从 Excel 字节流批量导入学生账号。

        Args:
            file_bytes: Excel 文件的原始字节内容。
            default_password: 密码列为空时使用的默认密码，仍为空则由 random_pwd_fn 生成。
            password_hasher: 将明文密码转换为哈希值的函数。
            random_pwd_fn: 生成随机密码的函数。

        Returns:
            含 total / success / skipped / failed / errors 字段的汇总 dict。

        Raises:
            ValueError: Excel 文件格式无效或解析失败。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, skipped, failed = 0, 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            display_name = str(row[0]).strip() if row[0] else ""
            student_id = str(row[1]).strip() if row[1] else ""
            grade = str(row[2]).strip() if row[2] else ""
            major = str(row[3]).strip() if row[3] else ""
            class_name = str(row[4]).strip() if row[4] else ""
            password = str(row[5]).strip() if len(row) > 5 and row[5] else (default_password or random_pwd_fn())

            if not student_id:
                failed.append({"row": row_idx, "student_id": student_id, "reason": "学号不能为空"})
                continue

            if self._user_store.get_user_by_student_id(student_id):
                skipped += 1
                continue

            try:
                user = self._user_store.create_user(
                    username=student_id,
                    hashed_pwd=password_hasher(password),
                    display_name=display_name,
                    role="student",
                )
                self._user_store.upsert_student_profile(user["id"], student_id, grade, major, class_name)
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] student import row %d failed: %s", row_idx, e)
                failed.append({"row": row_idx, "student_id": student_id, "reason": str(e)})

        return {
            "total": success + skipped + len(failed),
            "success": success,
            "skipped": skipped,
            "failed": len(failed),
            "errors": failed,
        }

    def import_teachers_from_xlsx(
        self,
        file_bytes: bytes,
        default_password: str,
        password_hasher: "callable[[str], str]",
        random_pwd_fn: "callable[[], str]",
    ) -> dict:
        """从 Excel 字节流批量导入教师账号。

        Args:
            file_bytes: Excel 文件的原始字节内容。
            default_password: 密码列为空时使用的默认密码，仍为空则由 random_pwd_fn 生成。
            password_hasher: 将明文密码转换为哈希值的函数。
            random_pwd_fn: 生成随机密码的函数。

        Returns:
            含 total / success / skipped / failed / errors 字段的汇总 dict。

        Raises:
            ValueError: Excel 文件格式无效或解析失败。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, skipped, failed = 0, 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            display_name, emp_id, dept, title = [str(x).strip() if x else "" for x in row[:4]]
            password = str(row[4]).strip() if len(row) > 4 and row[4] else (default_password or random_pwd_fn())

            if not emp_id:
                failed.append({"row": row_idx, "employee_id": emp_id, "reason": "工号不能为空"})
                continue

            if self._user_store.get_user_by_employee_id(emp_id):
                skipped += 1
                continue

            try:
                user = self._user_store.create_user(emp_id, password_hasher(password), display_name, "teacher")
                self._user_store.upsert_teacher_profile(user["id"], emp_id, dept, title)
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] teacher import row %d failed: %s", row_idx, e)
                failed.append({"row": row_idx, "employee_id": emp_id, "reason": str(e)})

        return {
            "total": success + skipped + len(failed),
            "success": success,
            "skipped": skipped,
            "failed": len(failed),
            "errors": failed,
        }

    def import_mentor_relations_from_xlsx(self, file_bytes: bytes) -> dict:
        """从 Excel 字节流批量导入师生关系。

        Args:
            file_bytes: Excel 文件的原始字节内容（学号列 + 工号列）。

        Returns:
            含 total / success / failed / errors 字段的汇总 dict。

        Raises:
            ValueError: Excel 文件格式无效或解析失败。
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except (ValueError, OSError, zipfile.BadZipFile) as e:
            raise ValueError(f"Excel 解析失败：{e}") from e

        success, failed = 0, []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            student_id = str(row[0]).strip() if row[0] else ""
            employee_id = str(row[1]).strip() if row[1] else ""

            if not student_id or not employee_id:
                failed.append({"row": row_idx, "reason": "学号和工号均不能为空"})
                continue

            student_user = self._user_store.get_user_by_student_id(student_id)
            if not student_user:
                failed.append({"row": row_idx, "student_id": student_id, "reason": "学生不存在"})
                continue

            mentor_user = self._user_store.get_user_by_employee_id(employee_id)
            if not mentor_user or mentor_user["role"] != "teacher":
                failed.append(
                    {
                        "row": row_idx,
                        "employee_id": employee_id,
                        "reason": "导师不存在或非教师角色",
                    }
                )
                continue

            try:
                self._user_store.add_mentor_relation(mentor_user["id"], student_user["id"])
                success += 1
            except (pymysql.Error, StorageError) as e:
                self.logger.warning("[UserService] relation import row %d failed: %s", row_idx, e)
                failed.append(
                    {
                        "row": row_idx,
                        "student_id": student_id,
                        "employee_id": employee_id,
                        "reason": str(e),
                    }
                )

        return {
            "total": success + len(failed),
            "success": success,
            "failed": len(failed),
            "errors": failed,
        }

    def export_students(self) -> list[dict]:
        """导出所有学生账号及档案。

        Returns:
            含展示字段的学生信息 dict 列表，适合直接传给 build_student_workbook。
        """
        items, _ = self._user_store.list_users(role="student", page=1, page_size=10000)
        rows = []
        for u in items:
            profile = self._user_store.get_student_profile(u["id"]) or {}
            rows.append(
                {
                    "display_name": u["display_name"],
                    "student_id": profile.get("student_id", ""),
                    "grade": profile.get("grade", ""),
                    "major": profile.get("major", ""),
                    "class_name": profile.get("class_name", ""),
                }
            )
        return rows

    def export_teachers(self) -> list[dict]:
        """导出所有教师账号及档案。

        Returns:
            含展示字段的教师信息 dict 列表，适合直接传给 build_teacher_workbook。
        """
        items, _ = self._user_store.list_users(role="teacher", page=1, page_size=10000)
        rows = []
        for u in items:
            profile = self._user_store.get_teacher_profile(u["id"]) or {}
            rows.append(
                {
                    "display_name": u["display_name"],
                    "employee_id": profile.get("employee_id", ""),
                    "department": profile.get("department", ""),
                    "title": profile.get("title", ""),
                }
            )
        return rows
