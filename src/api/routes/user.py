"""用户管理路由（管理员/教师用）：学生账号 CRUD、档案管理。"""

import io
import logging
from datetime import date

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from openpyxl import load_workbook

from src.api.auth import (
    get_current_user,
    hash_password,
    require_admin,
    require_teacher_or_admin,
)
from src.api.schemas import (
    MentorRelationRequest,
    MessageResponse,
    PaginatedUsers,
    ResetPasswordRequest,
    UpdateProfileRequest,
    UserCreate,
    UserInfo,
    UserUpdate,
)
from src.core.user_import import (
    build_relations_template_workbook,
    build_student_workbook,
    build_teacher_workbook,
    make_xlsx_response,
    random_password,
)
from src.storage.user_store import UserStore

router = APIRouter(prefix="/api/users", tags=["users"])
logger = logging.getLogger(__name__)

_us = UserStore()


def _to_user_info(user: dict, profile: dict | None = None) -> dict:
    return {
        "id": user["id"],
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "is_active": bool(user["is_active"]),
        "created_at": user["created_at"],
        "updated_at": user["updated_at"],
        "profile": profile,
    }


def _get_profile(user: dict) -> dict | None:
    role = user["role"]
    uid = user["id"]
    if role == "student":
        return _us.get_student_profile(uid)
    if role in ("teacher", "admin"):
        return _us.get_teacher_profile(uid)
    return None


@router.get("", response_model=PaginatedUsers)
def list_users(
    role: str | None = Query(default=None, pattern=r"^(admin|teacher|student)$"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    current_user: dict = Depends(require_teacher_or_admin),
):
    """列出用户（分页）。教师只能看学生；管理员看所有。"""
    effective_role = role
    if current_user["role"] == "teacher":
        if role not in (None, "student"):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="教师只能查看学生账号")
        effective_role = "student"

    items, total = _us.list_users(role=effective_role, page=page, page_size=page_size)
    return {
        "items": [_to_user_info(u, _get_profile(u)) for u in items],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("", response_model=UserInfo, status_code=status.HTTP_201_CREATED)
def create_user(body: UserCreate, current_user: dict = Depends(require_admin)):
    """创建用户（管理员专用）。"""
    if _us.get_user_by_username(body.username):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"用户名 '{body.username}' 已存在",
        )
    user = _us.create_user(
        username=body.username,
        hashed_pwd=hash_password(body.password),
        display_name=body.display_name,
        role=body.role,
    )
    return _to_user_info(user)


@router.get("/{user_id}", response_model=UserInfo)
def get_user(user_id: int, current_user: dict = Depends(require_teacher_or_admin)):
    """获取用户详情（含档案）。"""
    user = _us.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    return _to_user_info(user, _get_profile(user))


@router.put("/{user_id}", response_model=UserInfo)
def update_user(user_id: int, body: UserUpdate, current_user: dict = Depends(require_admin)):
    """更新用户基本信息（管理员专用）。"""
    user = _us.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    updates = body.model_dump(exclude_none=True)
    if updates:
        user = _us.update_user(user_id, **updates)
    return _to_user_info(user, _get_profile(user))


@router.put("/{user_id}/profile")
def update_profile(
    user_id: int,
    body: UpdateProfileRequest,
    current_user: dict = Depends(require_teacher_or_admin),
):
    """更新用户档案（学生档案或教师档案）。"""
    user = _us.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    if user["role"] == "student" and body.student_profile:
        sp = body.student_profile
        profile = _us.upsert_student_profile(user_id, sp.student_id, sp.grade, sp.major, sp.class_name)
        return {"message": "学生档案更新成功", "profile": profile}

    if user["role"] in ("teacher", "admin") and body.teacher_profile:
        tp = body.teacher_profile
        profile = _us.upsert_teacher_profile(user_id, tp.employee_id, tp.department, tp.title)
        return {"message": "教师档案更新成功", "profile": profile}

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="档案类型与用户角色不匹配")


@router.delete("/{user_id}", response_model=MessageResponse)
def delete_user(user_id: int, current_user: dict = Depends(require_admin)):
    """删除用户（管理员专用，不能删除自己）。"""
    if user_id == current_user["id"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不能删除自己的账号")
    user = _us.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    _us.delete_user(user_id)
    return {"message": f"用户 '{user['username']}' 已删除"}


@router.put("/{user_id}/reset-password", response_model=MessageResponse)
def reset_password(
    user_id: int,
    body: ResetPasswordRequest,
    current_user: dict = Depends(require_admin),
):
    """重置用户密码（管理员专用）。"""
    user = _us.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    _us.update_user(user_id, hashed_pwd=hash_password(body.new_password))
    return {"message": "密码重置成功"}


# ── Excel 导入/导出 ───────────────────────────────────────────


@router.get("/students/template")
def download_student_template(current_user: dict = Depends(require_teacher_or_admin)):
    """下载学生批量导入模板。"""
    wb = build_student_workbook(rows=None)
    return make_xlsx_response(wb, "学生账号导入模板.xlsx")


@router.get("/students/export")
def export_students_excel(current_user: dict = Depends(require_teacher_or_admin)):
    """导出所有学生账号为 Excel。"""
    items, _ = _us.list_users(role="student", page=1, page_size=10000)
    rows = []
    for u in items:
        profile = _us.get_student_profile(u["id"]) or {}
        rows.append(
            {
                "display_name": u["display_name"],
                "student_id": profile.get("student_id", ""),
                "grade": profile.get("grade", ""),
                "major": profile.get("major", ""),
                "class_name": profile.get("class_name", ""),
            }
        )
    filename = f"学生账号_{date.today().strftime('%Y%m%d')}.xlsx"
    return make_xlsx_response(build_student_workbook(rows), filename)


@router.post("/students/import")
async def import_students_excel(
    file: UploadFile = File(...),
    default_password: str = Form(default=""),
    current_user: dict = Depends(require_admin),
) -> dict:
    """从 Excel 批量导入学生账号（管理员专用）。

    - 用户名或学号已存在的行自动跳过
    - 密码列留空时使用 default_password 参数，仍为空则随机生成 10 位
    """
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件过大，请控制在 5MB 以内")

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 解析失败：{e}") from e

    success, skipped, failed = 0, 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        display_name = str(row[0]).strip() if row[0] else ""
        student_id = str(row[1]).strip() if row[1] else ""
        grade = str(row[2]).strip() if row[2] else ""
        major = str(row[3]).strip() if row[3] else ""
        class_name = str(row[4]).strip() if row[4] else ""
        password = str(row[5]).strip() if len(row) > 5 and row[5] else (default_password or random_password())

        if not student_id:
            failed.append({"row": row_idx, "student_id": student_id, "reason": "学号不能为空"})
            continue

        if _us.get_user_by_student_id(student_id):
            skipped += 1
            continue

        try:
            user = _us.create_user(
                username=student_id,
                hashed_pwd=hash_password(password),
                display_name=display_name,
                role="student",
            )
            _us.upsert_student_profile(user["id"], student_id, grade, major, class_name)
            success += 1
        except Exception as e:
            logger.warning("[student import] 第 %d 行导入失败: %s", row_idx, e)
            failed.append({"row": row_idx, "student_id": student_id, "reason": str(e)})

    return {
        "total": success + skipped + len(failed),
        "success": success,
        "skipped": skipped,
        "failed": len(failed),
        "errors": failed,
    }


@router.get("/mentors/relations/template")
def download_relations_template(current_user: dict = Depends(require_teacher_or_admin)):
    """下载师生关系批量导入模板。"""
    wb = build_relations_template_workbook()
    return make_xlsx_response(wb, "师生关系导入模板.xlsx")


@router.post("/mentors/relations/import")
async def import_mentor_relations(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin),
) -> dict:
    """从 Excel 批量导入师生关系（管理员专用）。"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 解析失败：{e}") from e

    success, failed = 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        student_id = str(row[0]).strip() if row[0] else ""
        employee_id = str(row[1]).strip() if row[1] else ""

        if not student_id or not employee_id:
            failed.append({"row": row_idx, "reason": "学号和工号均不能为空"})
            continue

        student_user = _us.get_user_by_student_id(student_id)
        if not student_user:
            failed.append({"row": row_idx, "student_id": student_id, "reason": "学生不存在"})
            continue

        mentor_user = _us.get_user_by_employee_id(employee_id)
        if not mentor_user or mentor_user["role"] != "teacher":
            failed.append(
                {
                    "row": row_idx,
                    "employee_id": employee_id,
                    "reason": "导师不存在或非教师角色",
                }
            )
            continue

        try:
            _us.add_mentor_relation(mentor_user["id"], student_user["id"])
            success += 1
        except Exception as e:
            logger.warning("[relation import] 第 %d 行导入失败: %s", row_idx, e)
            failed.append(
                {
                    "row": row_idx,
                    "student_id": student_id,
                    "employee_id": employee_id,
                    "reason": str(e),
                }
            )

    return {
        "total": success + len(failed),
        "success": success,
        "failed": len(failed),
        "errors": failed,
    }


# ── 教师账号 Excel 导入/导出 ──────────────────────────────────


@router.get("/teachers/template")
def download_teacher_template(current_user: dict = Depends(require_admin)):
    """下载教师批量导入模板。"""
    wb = build_teacher_workbook(rows=None)
    return make_xlsx_response(wb, "教师账号导入模板.xlsx")


@router.get("/teachers/export")
def export_teachers_excel(current_user: dict = Depends(require_admin)):
    """导出所有教师账号。"""
    items, _ = _us.list_users(role="teacher", page=1, page_size=10000)
    rows = []
    for u in items:
        profile = _us.get_teacher_profile(u["id"]) or {}
        rows.append(
            {
                "display_name": u["display_name"],
                "employee_id": profile.get("employee_id", ""),
                "department": profile.get("department", ""),
                "title": profile.get("title", ""),
            }
        )
    filename = f"教师账号_{date.today().strftime('%Y%m%d')}.xlsx"
    return make_xlsx_response(build_teacher_workbook(rows), filename)


@router.post("/teachers/import")
async def import_teachers_excel(
    file: UploadFile = File(...),
    default_password: str = Form(default=""),
    current_user: dict = Depends(require_admin),
) -> dict:
    """从 Excel 批量导入教师账号（管理员专用）。"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 解析失败：{e}") from e

    success, skipped, failed = 0, 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        display_name, emp_id, dept, title = [str(x).strip() if x else "" for x in row[:4]]
        password = str(row[4]).strip() if len(row) > 4 and row[4] else (default_password or random_password())

        if not emp_id:
            failed.append({"row": row_idx, "employee_id": emp_id, "reason": "工号不能为空"})
            continue
        if _us.get_user_by_employee_id(emp_id):
            skipped += 1
            continue

        try:
            user = _us.create_user(emp_id, hash_password(password), display_name, "teacher")
            _us.upsert_teacher_profile(user["id"], emp_id, dept, title)
            success += 1
        except Exception as e:
            failed.append({"row": row_idx, "employee_id": emp_id, "reason": str(e)})

    return {
        "total": success + skipped + len(failed),
        "success": success,
        "skipped": skipped,
        "failed": len(failed),
        "errors": failed,
    }


# ── 师生关系管理 ──────────────────────────────────────────────


@router.get("/mentors/{mentor_id}/students", response_model=list[UserInfo])
def list_mentor_students(mentor_id: int, current_user: dict = Depends(require_teacher_or_admin)):
    """列出指定导师名下的学生。"""
    if current_user["role"] == "teacher" and current_user["id"] != mentor_id:
        raise HTTPException(status_code=403, detail="无权查看其他导师的学生")

    students = _us.list_mentor_students(mentor_id)
    return [
        _to_user_info(
            u,
            {
                "student_id": u["student_id"],
                "grade": u["grade"],
                "major": u["major"],
                "class_name": u["class_name"],
            },
        )
        for u in students
    ]


@router.post("/mentors/relations", response_model=MessageResponse)
def add_mentor_relations(body: MentorRelationRequest, current_user: dict = Depends(require_admin)):
    """批量绑定师生关系（管理员专用）。"""
    mentor = _us.get_user_by_id(body.mentor_id)
    if not mentor or mentor["role"] != "teacher":
        raise HTTPException(status_code=400, detail="指定的导师 ID 无效或非教师角色")

    for sid in body.student_ids:
        _us.add_mentor_relation(body.mentor_id, sid)
    return {"message": f"成功为导师 {mentor['display_name']} 绑定 {len(body.student_ids)} 名学生"}


@router.delete("/mentors/{mentor_id}/students/{student_id}", response_model=MessageResponse)
def remove_mentor_relation(mentor_id: int, student_id: int, current_user: dict = Depends(require_admin)):
    """解除师生关系（管理员专用）。"""
    _us.remove_mentor_relation(mentor_id, student_id)
    return {"message": "解绑成功"}


@router.get("/me/mentor", response_model=UserInfo)
def get_my_mentor(current_user: dict = Depends(get_current_user)):
    """获取我的指导教师（学生专用）。"""
    if current_user["role"] != "student":
        raise HTTPException(status_code=400, detail="该接口仅供学生使用")

    mentor = _us.get_student_mentor(current_user["id"])
    if not mentor:
        raise HTTPException(status_code=404, detail="您尚未分配指导教师")

    return _to_user_info(
        mentor,
        {
            "employee_id": mentor["employee_id"],
            "department": mentor["department"],
            "title": mentor["title"],
        },
    )
