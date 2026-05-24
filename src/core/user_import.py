"""用户导入导出业务逻辑：Excel 工作簿构建、密码生成。

从 api/routes/user.py 提取，路由层仅负责参数解析和响应。
"""

import io
import secrets
import string
import urllib.parse

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── 学生工作簿 ─────────────────────────────────────────────────

_STU_COLS = ["姓名", "学号*", "年级", "专业", "班级", "初始密码（留空自动生成）"]
_STU_COL_WIDTH = [14, 18, 10, 22, 14, 28]


def build_student_workbook(rows: list[dict] | None) -> Workbook:
    """构建学生账号 Excel 工作簿。

    Args:
        rows: None → 模板（含示例行）；list → 导出（含数据）。

    Returns:
        openpyxl Workbook 实例。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "学生账号"

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, col in enumerate(_STU_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = _STU_COL_WIDTH[ci - 1]
    ws.row_dimensions[1].height = 22

    if rows is None:
        ws.append(["张三", "202201001", "2022", "计算机科学与技术", "计科一班", ""])
    else:
        for r in rows:
            ws.append(
                [
                    r.get("display_name", ""),
                    r.get("student_id", ""),
                    r.get("grade", ""),
                    r.get("major", ""),
                    r.get("class_name", ""),
                    "",
                ]
            )

    return wb


# ── 教师工作簿 ─────────────────────────────────────────────────

_TCH_COLS = ["姓名", "工号*", "院系", "职称", "初始密码（留空自动生成）"]
_TCH_COL_WIDTH = [14, 18, 22, 18, 28]


def build_teacher_workbook(rows: list[dict] | None) -> Workbook:
    """构建教师账号 Excel 工作簿。

    Args:
        rows: None → 模板（含示例行）；list → 导出（含数据）。

    Returns:
        openpyxl Workbook 实例。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "教师账号"

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, col in enumerate(_TCH_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = _TCH_COL_WIDTH[ci - 1]

    if rows is None:
        ws.append(["张老师", "T001", "计算机学院", "副教授", ""])
    else:
        for r in rows:
            ws.append(
                [
                    r.get("display_name", ""),
                    r.get("employee_id", ""),
                    r.get("department", ""),
                    r.get("title", ""),
                    "",
                ]
            )
    return wb


# ── 师生关系模板 ───────────────────────────────────────────────


def build_relations_template_workbook() -> Workbook:
    """构建师生关系批量导入模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "师生关系"

    cols = ["学生学号*", "导师工号*"]
    widths = [20, 20]

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, (col, width) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.append(["202201001", "T001"])
    return wb


# ── 工具函数 ───────────────────────────────────────────────────


def random_password(length: int = 10) -> str:
    """生成随机密码（字母+数字）。"""
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def make_xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    """将 openpyxl Workbook 转为 StreamingResponse。"""
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    encoded = urllib.parse.quote(filename)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
